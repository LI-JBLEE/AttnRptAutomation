"""
FY26 Attainment Club - Manager Report Generator
Generates per-manager attainment reports with hierarchy structure.
"""

import pandas as pd
import os
import re
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ── Configuration ──────────────────────────────────────────────────────────────
SOURCE_FILE = r"c:\Claude\Attainment report automation\FY26_Global_Attainment_Club.xlsx"
OUTPUT_DIR = r"c:\Claude\Attainment report automation\Manager report"
REPORT_DATE = datetime.today().strftime("%Y%m%d")

REPORT_COLUMNS = [
    "LI_EMP_ID", "Person Name", "Employee Status", "Level Grouping", "Level",
    "Fiscal Year", "Region", "Country", "Business_Unit", "Measure",
    "Plan_Period", "Level_1_Manager", "Level_2_Manager",
    "Q1 Credits", "Q1 Quota", "Q1 Att",
    "Q2 Credits", "Q2 Quota", "Q2 Att",
    "1H Credits", "1H Quota", "1H Att",
    "Q3 Credits", "Q3 Quota", "Q3 Att",
    "Q4 Credits", "Q4 Quota", "Q4 Att",
    "2H Credits", "2H Quota", "2H Att",
    "Annual Credits", "Annual Quota", "Annual Att",
    "Quota Start Date", "Quota End Date", "Measure Weight",
]

# Columns that contain attainment percentages
ATT_COLUMNS = ["Q1 Att", "Q2 Att", "1H Att", "Q3 Att", "Q4 Att", "2H Att", "Annual Att"]

# Columns that contain currency/number values
NUMBER_COLUMNS = [
    "Q1 Credits", "Q1 Quota", "Q2 Credits", "Q2 Quota",
    "1H Credits", "1H Quota", "Q3 Credits", "Q3 Quota",
    "Q4 Credits", "Q4 Quota", "2H Credits", "2H Quota",
    "Annual Credits", "Annual Quota",
]

# ── Styles ─────────────────────────────────────────────────────────────────────
COLORS = {
    "header_bg": "1B3A5C",       # Dark navy blue
    "header_font": "FFFFFF",     # White
    "border": "B0B0B0",         # Gray border
    "title_bg": "0F2B45",       # Very dark navy for title
    "title_font": "FFFFFF",
    "att_green": "27AE60",      # Green for >= 100%
    "att_yellow": "F39C12",     # Yellow/amber for 80-99%
    "att_red": "E74C3C",        # Red for < 80%
}

# Alternating level colors for hierarchy depth (data rows)
LEVEL_FILLS = [
    PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),  # L0: Light blue
    PatternFill(start_color="E8F0E8", end_color="E8F0E8", fill_type="solid"),  # L1: Light green
    PatternFill(start_color="F5E6F0", end_color="F5E6F0", fill_type="solid"),  # L2: Light purple
    PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),  # L3: Light orange
    PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid"),  # L4: Light cyan
    PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid"),  # L5: Light lime
    PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),  # L6: Light pink
]

# Section header colors per depth
SECTION_FILLS = [
    PatternFill(start_color="F0E6D3", end_color="F0E6D3", fill_type="solid"),  # L0: Warm beige
    PatternFill(start_color="E3D5C1", end_color="E3D5C1", fill_type="solid"),  # L1: Darker beige
    PatternFill(start_color="D6C8B3", end_color="D6C8B3", fill_type="solid"),  # L2: Even darker
    PatternFill(start_color="CBBDA8", end_color="CBBDA8", fill_type="solid"),  # L3
    PatternFill(start_color="C1B29D", end_color="C1B29D", fill_type="solid"),  # L4
    PatternFill(start_color="B8A894", end_color="B8A894", fill_type="solid"),  # L5
    PatternFill(start_color="AF9E8B", end_color="AF9E8B", fill_type="solid"),  # L6
]

thin_border = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

header_font = Font(name="Calibri", bold=True, size=10, color=COLORS["header_font"])
header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

title_font = Font(name="Calibri", bold=True, size=14, color=COLORS["title_font"])
title_fill = PatternFill(start_color=COLORS["title_bg"], end_color=COLORS["title_bg"], fill_type="solid")

subtitle_font = Font(name="Calibri", bold=False, size=10, color="666666")

section_font = Font(name="Calibri", bold=True, size=10, color="333333")

data_font = Font(name="Calibri", size=10)
data_alignment = Alignment(vertical="center")
data_alignment_center = Alignment(horizontal="center", vertical="center")
data_alignment_right = Alignment(horizontal="right", vertical="center")

# Column widths (tuned per content)
COLUMN_WIDTHS = {
    "LI_EMP_ID": 12, "Person Name": 28, "Employee Status": 14, "Level Grouping": 14,
    "Level": 10, "Fiscal Year": 10, "Region": 10, "Country": 14, "Business_Unit": 13,
    "Measure": 22, "Plan_Period": 14, "Level_1_Manager": 26, "Level_2_Manager": 26,
    "Q1 Credits": 13, "Q1 Quota": 13, "Q1 Att": 10,
    "Q2 Credits": 13, "Q2 Quota": 13, "Q2 Att": 10,
    "1H Credits": 13, "1H Quota": 13, "1H Att": 10,
    "Q3 Credits": 13, "Q3 Quota": 13, "Q3 Att": 10,
    "Q4 Credits": 13, "Q4 Quota": 13, "Q4 Att": 10,
    "2H Credits": 13, "2H Quota": 13, "2H Att": 10,
    "Annual Credits": 14, "Annual Quota": 14, "Annual Att": 11,
    "Quota Start Date": 15, "Quota End Date": 15, "Measure Weight": 14,
}


def get_att_font(value):
    """Return colored font based on attainment value."""
    if value is None or (isinstance(value, float) and pd.isna(value)) or value == 0:
        return Font(name="Calibri", size=10, color="999999")
    if value >= 1.0:
        return Font(name="Calibri", size=10, bold=True, color=COLORS["att_green"])
    elif value >= 0.8:
        return Font(name="Calibri", size=10, color=COLORS["att_yellow"])
    else:
        return Font(name="Calibri", size=10, color=COLORS["att_red"])


def extract_manager_name(full_name):
    """Extract clean name from 'First Last (ID)' format."""
    if pd.isna(full_name):
        return "Unknown"
    idx = full_name.rfind("(")
    if idx > 0:
        return full_name[:idx].strip()
    return full_name.strip()


def extract_manager_id(full_name):
    """Extract employee ID from 'First Last (ID)' format."""
    if pd.isna(full_name):
        return None
    start = full_name.rfind("(")
    end = full_name.rfind(")")
    if start > 0 and end > start:
        return full_name[start + 1:end].strip()
    return None


def sanitize_filename(name):
    """Remove characters not allowed in Windows filenames and strip parenthesized portions."""
    import re
    # Remove any parenthesized content (e.g. Chinese name aliases like "(黄策)")
    name = re.sub(r'\s*\([^)]*\)', '', name)
    # Remove illegal Windows filename characters
    for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
        name = name.replace(ch, '_')
    return name.strip()


def build_id_mappings(df):
    """
    Build ID-based lookup tables to handle name mismatches between
    Person Name and Level_1_Manager columns.

    Returns:
        person_id_to_name: {emp_id: Person Name}  (from employee records)
        l1_mgr_ids: set of employee IDs that appear as L1 managers
        person_id_to_l1_mgr_name: {emp_id: Level_1_Manager string}
    """
    # Map employee ID -> Person Name (from employee data rows)
    person_id_to_name = {}
    for _, row in df.drop_duplicates(subset=["Person Name"]).iterrows():
        pid = extract_manager_id(row["Person Name"])
        if pid:
            person_id_to_name[pid] = row["Person Name"]

    # Map employee ID -> Level_1_Manager string (the name used in L1_Manager column)
    person_id_to_l1_mgr_name = {}
    for mgr in df["Level_1_Manager"].dropna().unique():
        mid = extract_manager_id(mgr)
        if mid:
            person_id_to_l1_mgr_name[mid] = mgr

    # Set of employee IDs that are L1 managers
    l1_mgr_ids = set(person_id_to_l1_mgr_name.keys())

    return person_id_to_name, l1_mgr_ids, person_id_to_l1_mgr_name


def build_manager_region_map(df, person_id_to_name):
    """
    Build a mapping of manager full name (L1_Manager string) -> Region.
    Uses ID-based matching to find the manager's employee record.
    """
    manager_region = {}
    l1_managers = df["Level_1_Manager"].dropna().unique()

    # Build ID -> Region from employee records
    id_to_region = {}
    for _, row in df.drop_duplicates(subset=["Person Name"]).iterrows():
        pid = extract_manager_id(row["Person Name"])
        if pid and pd.notna(row.get("Region")):
            id_to_region[pid] = row["Region"]

    for mgr in l1_managers:
        mid = extract_manager_id(mgr)
        # Pass 1: match by ID in employee records
        if mid and mid in id_to_region:
            manager_region[mgr] = id_to_region[mid]
        else:
            # Pass 2: infer from direct reports' most common Region
            reports = df[df["Level_1_Manager"] == mgr]
            if len(reports) > 0:
                mode = reports["Region"].mode()
                manager_region[mgr] = mode.iloc[0] if len(mode) > 0 else "OTHER"
            else:
                manager_region[mgr] = "OTHER"

    return manager_region


def build_hierarchy_data(df, manager_name, l1_mgr_ids, person_id_to_l1_mgr_name,
                         depth=0, visited=None):
    """
    Recursively build hierarchical data for a manager's report.
    Uses ID-based matching to detect sub-managers (handles name mismatches).

    Returns list of tuples:
      - Data row:    (depth, row_data)
      - Section hdr: (-1, "SECTION", section_name, depth)
    """
    if visited is None:
        visited = set()

    # Guard against circular references (by manager name string)
    if manager_name in visited:
        return []
    visited.add(manager_name)

    result = []

    # Get direct reports (people whose Level_1_Manager is this manager)
    direct_reports = df[df["Level_1_Manager"] == manager_name]

    # Separate direct reports into non-managers and managers using ID matching
    direct_persons = direct_reports["Person Name"].unique()
    direct_managers = []
    direct_non_managers = []

    for person in direct_persons:
        pid = extract_manager_id(person)
        if pid and pid in l1_mgr_ids:
            direct_managers.append(person)
        else:
            direct_non_managers.append(person)

    # First: add non-manager direct reports at current depth
    for person in sorted(direct_non_managers):
        person_rows = direct_reports[direct_reports["Person Name"] == person]
        for _, row in person_rows.iterrows():
            result.append((depth, row))

    # Then: for each direct report who is also a manager, recurse
    for mgr_person in sorted(direct_managers):
        # Add section header at current depth
        result.append((-1, "SECTION", mgr_person, depth))

        # Add the sub-manager's own data rows at current depth
        mgr_rows = direct_reports[direct_reports["Person Name"] == mgr_person]
        for _, row in mgr_rows.iterrows():
            result.append((depth, row))

        # Find the L1_Manager name string for this person (may differ from Person Name)
        pid = extract_manager_id(mgr_person)
        l1_mgr_name = person_id_to_l1_mgr_name.get(pid, mgr_person)

        # Recursively add the sub-manager's entire team at depth+1
        sub_items = build_hierarchy_data(
            df, l1_mgr_name, l1_mgr_ids, person_id_to_l1_mgr_name,
            depth + 1, visited.copy()
        )
        result.extend(sub_items)

    return result


def write_report(wb, manager_name, hierarchy_data, report_columns):
    """Write a professional formatted report with recursive grouping."""
    ws = wb.active
    ws.title = "Attainment Report"

    clean_name = extract_manager_name(manager_name)
    # Remove any remaining parenthesized aliases (e.g. Chinese names)
    clean_name = re.sub(r'\s*\([^)]*\)', '', clean_name).strip()
    num_cols = len(report_columns)

    # ── Row 1: Title bar ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=f"FY26 Attainment Report — {clean_name}")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    for c in range(2, num_cols + 1):
        ws.cell(row=1, column=c).fill = title_fill

    # ── Row 2: Subtitle ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sub_cell = ws.cell(row=2, column=1,
                       value=f"Report Date: {datetime.today().strftime('%B %d, %Y')}    |    Manager: {clean_name}")
    sub_cell.font = subtitle_font
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Row 3: blank spacer ──
    ws.row_dimensions[3].height = 6

    # ── Row 4: Column headers ──
    header_row = 4
    for col_idx, col_name in enumerate(report_columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 30

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(num_cols)}{header_row}"

    # ── First pass: write data rows and record grouping ranges ──
    current_row = header_row + 1
    # Stack to track section grouping: [(section_start_row, depth), ...]
    group_ranges = []  # [(start_row, end_row, outline_level), ...]
    section_stack = []  # [(start_row, depth), ...]

    for item in hierarchy_data:
        if item[0] == -1 and item[1] == "SECTION":
            section_depth = item[3]  # depth at which this section appears

            # Close any open sections at same or deeper depth
            while section_stack and section_stack[-1][1] >= section_depth:
                start_row_s, depth_s = section_stack.pop()
                # Group rows from start+1 (after section header) to current_row-1
                if current_row - 1 > start_row_s:
                    group_ranges.append((start_row_s + 1, current_row - 1, depth_s + 1))

            # Section header row
            section_name = extract_manager_name(item[2])
            indent_str = "  " * section_depth
            s_fill = SECTION_FILLS[min(section_depth, len(SECTION_FILLS) - 1)]

            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=num_cols)
            cell = ws.cell(row=current_row, column=1,
                           value=f"{indent_str}▸ Team: {section_name}")
            cell.font = section_font
            cell.fill = s_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            for c in range(2, num_cols + 1):
                sc = ws.cell(row=current_row, column=c)
                sc.fill = s_fill
                sc.border = thin_border
            ws.row_dimensions[current_row].height = 22

            # Push this section onto the stack
            section_stack.append((current_row, section_depth))
            current_row += 1
            continue

        # Data row
        level = item[0]
        row_data = item[1]
        row_fill = LEVEL_FILLS[min(level, len(LEVEL_FILLS) - 1)]
        indent = "  " * level

        for col_idx, col_name in enumerate(report_columns, 1):
            value = row_data.get(col_name, "")
            if pd.isna(value):
                value = None

            cell = ws.cell(row=current_row, column=col_idx)

            if col_name == "Person Name" and indent:
                cell.value = indent + str(value) if value else ""
            elif col_name in ATT_COLUMNS:
                cell.value = value
                if value is not None and value != 0:
                    cell.number_format = '0.0%'
                    cell.font = get_att_font(value)
                else:
                    cell.font = Font(name="Calibri", size=10, color="999999")
                cell.alignment = data_alignment_right
                cell.fill = row_fill
                cell.border = thin_border
                continue
            elif col_name in NUMBER_COLUMNS:
                cell.value = value
                if value is not None and value != 0:
                    cell.number_format = '#,##0'
                cell.alignment = data_alignment_right
            elif col_name == "Measure Weight":
                cell.value = value
                if value is not None:
                    cell.number_format = '0%'
                cell.alignment = data_alignment_center
            elif col_name in ("Quota Start Date", "Quota End Date"):
                cell.value = value
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = data_alignment_center
            elif col_name in ("LI_EMP_ID", "Fiscal Year"):
                cell.value = value
                cell.alignment = data_alignment_center
            elif col_name in ("Employee Status", "Level Grouping", "Level", "Region",
                               "Country", "Business_Unit", "Plan_Period"):
                cell.value = value
                cell.alignment = data_alignment_center
            else:
                cell.value = value
                cell.alignment = data_alignment

            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # Close any remaining open sections
    while section_stack:
        start_row_s, depth_s = section_stack.pop()
        if current_row - 1 > start_row_s:
            group_ranges.append((start_row_s + 1, current_row - 1, depth_s + 1))

    # ── Apply Excel outline grouping (collapse/expand) ──
    # Sort by outline level ascending so nested groups are set correctly
    group_ranges.sort(key=lambda x: x[2])
    for start_r, end_r, outline_lvl in group_ranges:
        # Excel supports outline levels 1-8
        if outline_lvl > 8:
            outline_lvl = 8
        ws.row_dimensions.group(start_r, end_r, outline_level=outline_lvl, hidden=False)

    # Outline settings: summary rows above detail (section header is above its group)
    ws.sheet_properties.outlinePr.summaryBelow = False

    # ── Set column widths ──
    for col_idx, col_name in enumerate(report_columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(col_name, 12)

    # ── Freeze panes ──
    ws.freeze_panes = f"C{header_row + 1}"

    # ── Print settings ──
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def get_all_regions(source_df):
    """
    Get all unique regions from the attainment data.
    Returns sorted list of region strings.
    """
    person_id_to_name, _, _ = build_id_mappings(source_df)
    manager_region = build_manager_region_map(source_df, person_id_to_name)
    return sorted(set(manager_region.values()))


def generate_all_reports(source_df, output_dir, progress_callback=None,
                         selected_regions=None):
    """
    Generate all manager reports from a pre-loaded DataFrame.

    Args:
        source_df: pandas DataFrame (attainment data with Plan_Period column renamed)
        output_dir: output directory path for reports
        progress_callback: optional callable(current, total, message) for UI updates
        selected_regions: optional list of region strings to filter by (None = all)

    Returns:
        dict with keys: total, region_counts, managers
              managers is list of (manager_full_name, region, safe_name, filepath)
    """
    # Build ID-based lookup tables for hierarchy detection
    person_id_to_name, l1_mgr_ids, person_id_to_l1_mgr_name = build_id_mappings(source_df)

    # Build manager -> Region mapping (ID-based)
    manager_region = build_manager_region_map(source_df, person_id_to_name)

    os.makedirs(output_dir, exist_ok=True)

    # Get unique L1 managers, optionally filtered by region
    all_managers = sorted(source_df["Level_1_Manager"].dropna().unique())
    if selected_regions is not None:
        region_set = set(selected_regions)
        managers = [m for m in all_managers if manager_region.get(m, "OTHER") in region_set]
    else:
        managers = all_managers

    # Clean up only the region subfolders that will be regenerated
    regions_to_generate = set(
        manager_region.get(m, "OTHER") for m in managers
    )
    for region in regions_to_generate:
        region_dir = os.path.join(output_dir, region)
        if os.path.exists(region_dir):
            # Only remove existing report files, not the entire folder tree
            for fname in os.listdir(region_dir):
                if fname.startswith("FY26_Attainment_") and fname.endswith(".xlsx"):
                    os.remove(os.path.join(region_dir, fname))

    total = len(managers)
    region_counts = {}
    report_date = datetime.today().strftime("%Y%m%d")
    generated_managers = []

    for i, manager in enumerate(managers, 1):
        clean_name = extract_manager_name(manager)
        safe_name = sanitize_filename(clean_name)
        filename = f"FY26_Attainment_{safe_name}_{report_date}.xlsx"

        # Determine Region subfolder
        region = manager_region.get(manager, "OTHER")
        region_dir = os.path.join(output_dir, region)
        os.makedirs(region_dir, exist_ok=True)
        filepath = os.path.join(region_dir, filename)

        region_counts[region] = region_counts.get(region, 0) + 1

        # Build hierarchy data (ID-based matching)
        hierarchy_data = build_hierarchy_data(
            source_df, manager, l1_mgr_ids, person_id_to_l1_mgr_name
        )

        if not hierarchy_data:
            continue

        # Create workbook and write report
        wb = Workbook()
        write_report(wb, manager, hierarchy_data, REPORT_COLUMNS)
        wb.save(filepath)

        generated_managers.append((manager, region, safe_name, filepath))

        if progress_callback:
            progress_callback(i, total, f"{region}/{filename}")

    return {
        "total": total,
        "region_counts": region_counts,
        "managers": generated_managers,
    }


def main():
    print("Loading source data...")
    df = pd.read_excel(SOURCE_FILE, sheet_name="in")

    # Rename the Plan_Period column
    df.rename(columns={"Plan_Period;MBO_Description": "Plan_Period"}, inplace=True)

    def cli_progress(current, total, message):
        if current % 50 == 0 or current == total:
            print(f"  [{current}/{total}] Generated: {message}")

    print("Generating reports...")
    results = generate_all_reports(df, OUTPUT_DIR, progress_callback=cli_progress)

    print(f"\nDone! {results['total']} reports saved to: {OUTPUT_DIR}")
    print("Region distribution:")
    for region, count in sorted(results["region_counts"].items()):
        print(f"  {region}: {count} reports")


if __name__ == "__main__":
    main()
